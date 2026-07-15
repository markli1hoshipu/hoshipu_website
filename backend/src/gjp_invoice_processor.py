"""
GJP 发票导出 — extract flight e-invoice PDFs (航空运输电子客票行程单)
into the base columns of the YUHANG SAP-reconciliation Excel format.

Only the columns that can be read off the invoice are produced here; the
derived columns (tax base, Net, tax, Des. strings, SAP date) are Excel
formulas written by the workbook builder so they auto-compute on open.

Two invoice layouts are handled:
  A) halfwidth colons, segment like  "自:上海 浦东 吉祥 HO1163 ..."  (city + airport, needs IATA lookup)
  B) fullwidth colons, segment like  "自：DLC 大连 南航 CZ6449 ..."   (IATA code inline)
"""
import io
import os
import re
from typing import Optional, Dict, Any, List

import pdfplumber

# ---------------------------------------------------------------------------
# Airport IATA lookup (城市/机场 -> IATA code)
# ---------------------------------------------------------------------------
_AIRPORTS_PATH = os.path.join(os.path.dirname(__file__), "assets", "airports.txt")

# Airports that appear on invoices but are missing from airports.txt.
_EXTRA_AIRPORTS: List[tuple] = [
    ("WEF", "潍坊华夏机场"),
]


def _load_airports() -> List[tuple]:
    """Return list of (IATA, chinese_name) in file order (order matters:
    the first match wins, so preferred codes must be listed first)."""
    pairs: List[tuple] = []
    try:
        with open(_AIRPORTS_PATH, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line == "END123":
                    continue
                parts = line.split(None, 1)
                if len(parts) == 2:
                    pairs.append((parts[0].strip(), parts[1].strip()))
    except FileNotFoundError:
        pass
    pairs.extend(_EXTRA_AIRPORTS)
    return pairs


_AIRPORTS = _load_airports()


def lookup_iata(city: str, airport: str) -> Optional[str]:
    """Resolve a Chinese city + airport name to an IATA code.

    The airport token (浦东/虹桥/首都/周水子) is the most distinctive, so it is
    tried first; the city is only a fallback. First match wins."""
    airport = (airport or "").strip()
    city = (city or "").strip()
    # 1) match on the distinctive airport token
    if airport:
        for iata, cn in _AIRPORTS:
            if airport in cn:
                return iata
    # 2) fall back to the combined city+airport being a prefix of the name
    combined = (city + airport)
    if combined:
        for iata, cn in _AIRPORTS:
            if cn.startswith(combined):
                return iata
    # 3) last resort: match on the city
    if city:
        for iata, cn in _AIRPORTS:
            if city in cn:
                return iata
    return None


# ---------------------------------------------------------------------------
# Regexes
# ---------------------------------------------------------------------------
_SEG_RE = re.compile(r"^(自|至)[：:]\s*(.*)$")
_FLIGHT_RE = re.compile(r"\b([A-Z0-9]{2}\d{3,4})\b")
_DATE_RE = re.compile(r"(\d{4})年(\d{1,2})月(\d{1,2})日")
_IATA_TOKEN_RE = re.compile(r"^[A-Z]{3}$")
_CNY_RE = re.compile(r"CNY\s*(-?[\d,]+\.?\d*)")


def _fmt_date(y: str, m: str, d: str) -> str:
    return f"{int(d):02d}.{int(m):02d}.{y}"


def _to_number(s: str) -> Optional[float]:
    try:
        return float(s.replace(",", ""))
    except (ValueError, AttributeError):
        return None


def extract_text_from_pdf(pdf_bytes: bytes) -> str:
    full_text = ""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                full_text += page_text + "\n"
    return full_text


def _parse_segments(text: str) -> List[Dict[str, Any]]:
    """Parse the 自:/至: flight legs into an ordered list of
    {city, iata, flight, date}. Empty 至: lines are skipped."""
    legs: List[Dict[str, Any]] = []
    for raw in text.splitlines():
        m = _SEG_RE.match(raw.strip())
        if not m:
            continue
        rest = m.group(2).strip()
        if not rest:
            continue
        toks = rest.split()
        if not toks:
            continue
        # In layout A the fare row is itself a "至: CNY 972.48 ..." line; skip it.
        if toks[0] == "CNY":
            continue

        if _IATA_TOKEN_RE.match(toks[0]):
            # Layout B: "DLC 大连 ..."
            iata = toks[0]
            city = toks[1] if len(toks) > 1 else ""
            after = toks[2:]
        else:
            # Layout A: "上海 浦东 ..."
            city = toks[0]
            airport = toks[1] if len(toks) > 1 else ""
            iata = lookup_iata(city, airport)
            after = toks[2:]

        fm = _FLIGHT_RE.search(" ".join(after))
        flight = fm.group(1) if fm else None
        dm = _DATE_RE.search(raw)
        date = _fmt_date(*dm.groups()) if dm else None

        legs.append({"city": city, "iata": iata, "flight": flight, "date": date})
    return legs


def _parse_fare_line(text: str) -> Dict[str, Optional[float]]:
    """The fare row lists CNY amounts in a fixed leading order:
    票价, 燃油附加费, 增值税税额, 民航发展基金, [其他税费], 合计.
    So M(民航基金)=values[3] and N(合计)=values[-1] regardless of whether
    the optional 其他税费 column is present."""
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if "票价" in line and "民航发展基金" in line and "合计" in line:
            # amounts are on this line or the following one
            for cand in (lines[i + 1] if i + 1 < len(lines) else "", line):
                amounts = [_to_number(x) for x in _CNY_RE.findall(cand)]
                amounts = [a for a in amounts if a is not None]
                if len(amounts) >= 5:
                    return {
                        "fare": amounts[0],
                        "fuel": amounts[1],
                        "vat": amounts[2],
                        "caac_fund": amounts[3],
                        "total": amounts[-1],
                    }
            break
    return {"fare": None, "fuel": None, "vat": None, "caac_fund": None, "total": None}


def extract_invoice(text: str) -> Dict[str, Any]:
    """Extract all base + bonus fields from one invoice's text."""
    dom = re.search(r"国内国际标识[：:]\s*(国内|国际)", text)
    inv = re.search(r"发票号码[：:]\s*([0-9A-Za-z]+)", text)
    issue = re.search(r"填开日期[：:]\s*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日)", text)
    buyer = re.search(r"购买方名称[：:]\s*(\S+)", text)
    ticket = re.search(r"电子客票号码[：:]\s*([0-9]+)", text)

    # passenger name: the token(s) before the (masked) ID number on the line
    # after the "旅客姓名 ..." header
    name = None
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if line.strip().startswith("旅客姓名"):
            if i + 1 < len(lines):
                nm = re.match(
                    r"^(.+?)\s+([0-9A-Za-z]*\*{2,}[0-9A-Za-z]*)", lines[i + 1].strip()
                )
                if nm:
                    name = nm.group(1).strip()
                else:
                    # no masked id found; take leading non-space chunk
                    name = lines[i + 1].strip().split("  ")[0].split(" ")[0] or None
            break

    fares = _parse_fare_line(text)
    legs = _parse_segments(text)

    cities = [lg["city"] for lg in legs if lg["city"]]
    iatas = [lg["iata"] for lg in legs if lg["iata"]]
    flights = [lg["flight"] for lg in legs if lg["flight"]]
    dates = [lg["date"] for lg in legs if lg["date"]]

    total = fares["total"]
    category = "Flight"
    if total is not None and total < 0:
        category = "Refund"

    return {
        # base columns for the Excel
        "category": category,                                   # C 类别
        "intl_dom": "Dom." if (dom and dom.group(1) == "国内") else ("Int." if dom else None),  # D
        "name": name,                                           # F 中文名
        "depart_date": dates[0] if dates else None,             # H 起飞时间
        "routing": "".join(iatas) if iatas else None,           # J ROUTING
        "route_cn": "-".join(cities) if cities else None,       # K 航程
        "flight_no": "/".join(flights) if flights else None,    # L 航班
        "caac_fund": fares["caac_fund"],                        # M 民航建设基金
        "total": total,                                         # N 合计金额
        "invoice_number": inv.group(1) if inv else None,        # T 发票号
        # bonus / diagnostic fields (not written to the sheet by default)
        "buyer": buyer.group(1) if buyer else None,
        "issue_date": issue.group(1) if issue else None,
        "ticket_number": ticket.group(1) if ticket else None,
        "fare": fares["fare"],
        "fuel_surcharge": fares["fuel"],
        "vat": fares["vat"],
        "unmapped_airports": [lg["city"] + (lg.get("airport") or "") for lg in legs if not lg["iata"]],
    }


# Base columns that must be present for a row to be considered complete.
REQUIRED_FIELDS = ["name", "intl_dom", "route_cn", "flight_no", "total", "invoice_number"]


def validate(info: Dict[str, Any]) -> Dict[str, Any]:
    missing = [f for f in REQUIRED_FIELDS if not info.get(f)]
    if info.get("routing") is None or (info.get("unmapped_airports")):
        missing.append("routing")
    return {"valid": len(missing) == 0, "missing": missing}


# ---------------------------------------------------------------------------
# Workbook builder — fill a copy of the template so the derived columns
# (I, W, X, Y, Z, AA, AB, AC, AE) auto-compute in Excel on open.
# ---------------------------------------------------------------------------
from copy import copy as _copy

_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "assets", "gjp_invoice_template.xlsx")
_DATA_SHEET = "电子发票"
_PROTOTYPE_ROW = 8

# base column letter -> key in the extracted info dict
_BASE_COL_MAP = {
    "C": "category",
    "D": "intl_dom",
    "F": "name",
    "H": "depart_date",
    "J": "routing",
    "K": "route_cn",
    "L": "flight_no",
    "M": "caac_fund",
    "N": "total",
    "T": "invoice_number",
    "U": "ticket_number",
}

# derived formula templates ({r} -> row number). Mirror the source workbook.
_DERIVED_FORMULAS = {
    "I": "=MID(H{r},1,2)&MID(H{r},4,2)&MID(H{r},9,2)",
    "W": '=IF(AND(OR(C{r}="Flight",C{r}="Train"),D{r}="Dom."),N{r}-M{r},0)',
    "X": "=VLOOKUP(C{r},GL!A:B,2,FALSE)",
    "Y": '=IF(Z{r}="j0",N{r},W{r}-AA{r}+M{r})',
    "Z": '=IF(AND(OR(C{r}="Flight",C{r}="Train"),D{r}="Dom."),"Y9","J0")',
    "AA": "=ROUND((W{r}/1.09)*0.09,2)",
    "AB": '=E{r}&L{r}&"_"&F{r}&"_"&J{r}&AC{r}',
    "AC": '=IF(OR(C{r}="rebook",C{r}="refund"),"_"&Q{r}&C{r},"_"&I{r})',
    "AE": '=IF(OR(C{r}="rebook",C{r}="refund"),H{r}," ")',
}

_MAX_COL = 33  # up to column AG, matching the template


def build_workbook(rows: List[Dict[str, Any]]) -> bytes:
    """Return .xlsx bytes: the template with one data row per extracted invoice."""
    from openpyxl import load_workbook

    wb = load_workbook(_TEMPLATE_PATH)
    ws = wb[_DATA_SHEET]

    for idx, info in enumerate(rows):
        r = _PROTOTYPE_ROW + idx

        # replicate the prototype row's styling for every new row
        if r != _PROTOTYPE_ROW:
            for c in range(1, _MAX_COL + 1):
                src = ws.cell(row=_PROTOTYPE_ROW, column=c)
                dst = ws.cell(row=r, column=c)
                if src.has_style:
                    dst._style = _copy(src._style)

        ws[f"A{r}"] = idx + 1              # 序号
        ws[f"P{r}"] = "邮件"               # 邮件 (constant on every row in source)

        for col, key in _BASE_COL_MAP.items():
            ws[f"{col}{r}"] = info.get(key)

        for col, tmpl in _DERIVED_FORMULAS.items():
            ws[f"{col}{r}"] = tmpl.format(r=r)

    # make sure the file opens at the top-left, not scrolled down
    from openpyxl.worksheet.views import Selection

    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
    wb.active = wb.sheetnames.index(_DATA_SHEET)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
