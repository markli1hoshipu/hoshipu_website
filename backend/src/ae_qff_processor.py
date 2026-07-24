"""
AE 欠条报表 updater — merge QFF daily-IOU files into the master AE ledger.

The AE ledger has one sheet per month (YYYYMM). Within a sheet, debtor rows are
grouped by 负责人 in column A; columns B=欠款人, C=余额(formula), D=当月发生额(formula),
E=上月余额, and F,G,H… are one column PER DAY (row 1 holds the Excel serial date).
A merge writes each QFF IOU's 初始金额 into [debtor row, that day's column]; the
formulas recompute balances/totals on open.

Design decisions (confirmed with user):
  * amount written = 初始金额 (QFF col B)
  * debtor matched by space-normalized name within the QFF group (A == "QFF")
  * unmatched debtors -> a new row auto-inserted into the QFF group
  * duplicate imports prevented via a hidden "_已导入欠条id" log of 欠单号

Row insertion is done with a formula-remap so the ledger's hardcoded summary
ranges (grand total, group subtotals, row 541/542) stay correct — openpyxl does
NOT adjust formulas on insert, so we rewrite every same-sheet reference ourselves.
"""
import io
import os
import re
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy

LOG_SHEET = "_已导入欠条id"
_LEGACY_LOG_SHEETS = ["_QFF_imported"]  # older name — auto-migrated to LOG_SHEET
_EXCEL_MAX_ROW = 1048576
_FIRST_DAY_COL = 6  # column F — day columns start here (A..E are 负责人/欠款人/余额/发生额/上月余额)
_EPOCH = datetime(1899, 12, 30)


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def load_xlsx(data: bytes):
    """Load an .xlsx even if it is misnamed .xls (bypass extension check)."""
    return openpyxl.load_workbook(io.BytesIO(data))


def _norm(name: Any) -> str:
    """Normalize a debtor name for matching: drop all whitespace."""
    return re.sub(r"\s+", "", str(name)) if name is not None else ""


def _serial_to_yymmdd(serial: Any) -> Optional[str]:
    # A date-formatted header round-trips through Excel as a datetime, not a serial.
    if isinstance(serial, datetime):
        return serial.strftime("%y%m%d")
    if not isinstance(serial, (int, float)):
        return None
    try:
        return (_EPOCH + timedelta(days=float(serial))).strftime("%y%m%d")
    except (ValueError, OverflowError):
        return None


def _yymmdd_to_serial(yymmdd: str) -> Optional[int]:
    """Inverse of _serial_to_yymmdd: 'YYMMDD' -> Excel serial (int). '' if invalid."""
    if not re.fullmatch(r"\d{6}", yymmdd or ""):
        return None
    yy, mm, dd = int(yymmdd[0:2]), int(yymmdd[2:4]), int(yymmdd[4:6])
    try:
        d = datetime(2000 + yy, mm, dd)
    except ValueError:
        return None
    return (d - _EPOCH).days


def _date_to_ym_sheet(yymmdd: str) -> str:
    """'260710' -> '202607'."""
    return "20" + yymmdd[0:2] + yymmdd[2:4]


# --------------------------------------------------------------------------- #
# QFF reading
# --------------------------------------------------------------------------- #
def read_qff(data: bytes) -> List[Dict[str, Any]]:
    """Parse one QFF workbook into a list of IOU dicts."""
    wb = load_xlsx(data)
    ws = wb.worksheets[0]
    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        date = ws.cell(r, 1).value          # A 日期 (YYMMDD)
        initial = ws.cell(r, 2).value       # B 初始金额
        remaining = ws.cell(r, 3).value     # C 剩余金额
        debtor = ws.cell(r, 4).value        # D 欠款人
        owner = ws.cell(r, 5).value         # E 负责人
        iou_no = ws.cell(r, 6).value        # F 欠单号 (may be blank — still imported)
        remark = ws.cell(r, 7).value        # G 备注
        if debtor is None:                  # need a debtor to place the row
            continue
        date_s = str(date).strip() if date is not None else ""
        # normalize a YYMMDD-ish date (drop any trailing .0 from float cells)
        if date_s.endswith(".0"):
            date_s = date_s[:-2]
        rows.append({
            "iou_no": str(iou_no).strip() if iou_no is not None else "",
            "date": date_s,
            "initial": initial,
            "remaining": remaining,
            "debtor": str(debtor).strip(),
            "owner": str(owner).strip() if owner else "QFF",
            "remark": remark,
        })
    return rows


# --------------------------------------------------------------------------- #
# formula remap (for row insertion)
# --------------------------------------------------------------------------- #
_REF_RE = re.compile(r"(?<![A-Za-z0-9_!$'])(\$?)([A-Z]{1,3})(\$?)(\d+)")


def _remap_formula(formula: str, remap) -> str:
    """Rewrite same-sheet A1 row references using remap(old_row)->new_row.
    References that are sheet-qualified (preceded by '!') are skipped by the
    negative lookbehind, so cross-sheet refs are left untouched."""
    def repl(m: "re.Match") -> str:
        col_abs, col, row_abs, row = m.groups()
        new_row = remap(int(row))
        return f"{col_abs}{col}{row_abs}{new_row}"

    return _REF_RE.sub(repl, formula)


def _insert_rows_with_remap(ws, insert_before: int, count: int) -> None:
    """Insert `count` rows before original row `insert_before`, then fix every
    formula in the sheet so hardcoded ranges/refs stay correct."""
    ws.insert_rows(insert_before, count)

    def remap(row: int) -> int:
        if row >= insert_before:
            return min(row + count, _EXCEL_MAX_ROW)
        return row

    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                cell.value = _remap_formula(v, remap)


# --------------------------------------------------------------------------- #
# ledger introspection
# --------------------------------------------------------------------------- #
def _build_day_columns(ws) -> Dict[str, int]:
    """Map YYMMDD -> column index, from the serial dates in row 1."""
    cols: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        yymmdd = _serial_to_yymmdd(ws.cell(1, c).value)
        if yymmdd:
            cols.setdefault(yymmdd, c)
    return cols


def _ensure_day_column(ws, yymmdd: str, day_cols: Dict[str, int]) -> Optional[int]:
    """Return the column for `yymmdd`, creating it if the month sheet lacks it.

    Day columns are Excel serials in row 1 starting at column F; the slots to the
    right are pre-provisioned (row 2 already holds the daily-total formula, and the
    per-debtor C=SUM(E:ZC)/D=SUM(F:ZE) ranges already span them). So creating a day
    is just dropping the serial into the next free slot — no row/column insertion,
    no formula remap, no effect on any total. Returns None only if the date is
    invalid. `day_cols` is updated in place."""
    if yymmdd in day_cols:
        return day_cols[yymmdd]
    serial = _yymmdd_to_serial(yymmdd)
    if serial is None:
        return None
    used = list(day_cols.values())
    target = (max(used) + 1) if used else _FIRST_DAY_COL
    while ws.cell(1, target).value is not None:  # never overwrite an existing header
        target += 1
    hdr = ws.cell(1, target)
    hdr.value = serial
    if used:  # inherit the date number-format from an existing day header
        src = ws.cell(1, min(used))
        if src.has_style:
            hdr._style = copy(src._style)
    else:  # fresh sheet — style from the template + a readable date format
        _copy_style(_proto().cell(1, 6), hdr)
        hdr.number_format = "m/d"
    tot = ws.cell(2, target)
    if not (isinstance(tot.value, str) and tot.value.startswith("=")):
        letter = get_column_letter(target)
        tot.value = f"=SUM({letter}$3:{letter}${_EXCEL_MAX_ROW})"
    day_cols[yymmdd] = target
    return target


def _build_debtor_rows(ws, owner: str) -> Dict[str, int]:
    """Map normalized debtor name -> row, for rows whose 负责人 (col A) == owner."""
    out: Dict[str, int] = {}
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value == owner:
            name = ws.cell(r, 2).value
            if name is not None:
                out.setdefault(_norm(name), r)
    return out


def _last_group_row(ws, owner: str) -> Optional[int]:
    """The last row whose 负责人 (col A) == owner — new debtors go just after it.
    Returns None if the owner group is not present in the sheet."""
    last: Optional[int] = None
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value == owner:
            last = r
    return last


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        s = ws.cell(src_row, c)
        d = ws.cell(dst_row, c)
        if s.has_style:
            d._style = copy(s._style)


def _family_prefix(name: Any) -> str:
    """Group key = the 负责人 code before any suffix. 'QFF-月中' -> 'QFF', 'WD-纸' -> 'WD'."""
    return str(name or "").split("-")[0].strip()


# ---- styling for freshly-created sheets (match the AE-欠条报表 example colors) ----
_STYLE_PROTO = None  # cached 'proto' worksheet: row1=header, row2=labels, row3=debtor, row4=subtotal


def _proto():
    """Lazy-load the style template (row-prototypes extracted from the example ledger)."""
    global _STYLE_PROTO
    if _STYLE_PROTO is None:
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "assets", "ae_style_template.xlsx")
        _STYLE_PROTO = openpyxl.load_workbook(path)["proto"]
    return _STYLE_PROTO


def _copy_style(src, dst) -> None:
    """Copy visual style across workbooks (attribute copy — _style indices don't
    transfer between workbooks, but Font/Fill/Border/Alignment objects do)."""
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format


def _style_debtor_row(ws, r: int, max_col: int) -> None:
    """Apply the example's debtor-row colors (per column; day columns reuse col F)."""
    pr = _proto()
    for c in range(1, max_col + 1):
        _copy_style(pr.cell(3, c if c <= 5 else 6), ws.cell(r, c))


def _style_subtotal_row(ws, r: int) -> None:
    pr = _proto()
    _copy_style(pr.cell(4, 1), ws.cell(r, 1))
    _copy_style(pr.cell(4, 2), ws.cell(r, 2))


def _grand_total_row(ws) -> int:
    """Row of the '当月欠款增减合计' label — new groups are inserted just above it.
    Falls back to max_row+1 if the label isn't found."""
    for r in range(ws.max_row, 2, -1):
        if str(ws.cell(r, 1).value or "").strip() == "当月欠款增减合计":
            return r
    return ws.max_row + 1


def _open_grand_total(ws) -> None:
    """Make the sheet's grand total range-proof and re-anchor day-column totals.

    * grand total C1/D1 -> open-ended =SUM(C3:C1048576) / =SUM(D3:...), so any
      added row/group is always counted.
    * each day column's row-2 total -> =SUM(col$3:col$1048576) (row insertions may
      have shifted the anchor).

    Subtotal rows are deliberately left as contiguous =SUM(Cx:Cy): the row-insert
    remap keeps those ranges correct, and — unlike SUMIF($A:$A,…) — a bounded
    single-column SUM neither self-references its own column (no circular ref) nor
    hits Excel's flaky full-column-criteria behavior (which returned 0 / triggered
    a file repair)."""
    for col in (3, 4):  # C1 (余额合计), D1 (发生额合计)
        v = ws.cell(1, col).value
        if isinstance(v, str) and v.startswith("=SUM("):
            letter = get_column_letter(col)
            ws.cell(1, col).value = f"=SUM({letter}3:{letter}{_EXCEL_MAX_ROW})"
    for c in range(_FIRST_DAY_COL, ws.max_column + 1):
        if _serial_to_yymmdd(ws.cell(1, c).value):
            letter = get_column_letter(c)
            ws.cell(2, c).value = f"=SUM({letter}$3:{letter}${_EXCEL_MAX_ROW})"

    # extend 当月欠款增减合计 = SUM(B3:B{row-1}) so newly-inserted groups get counted
    # (keeps the column-B mechanism — its value differs from the raw D total by design).
    for r in range(ws.max_row, 2, -1):
        if str(ws.cell(r, 1).value or "").strip() == "当月欠款增减合计":
            b = ws.cell(r, 2).value
            if isinstance(b, str) and re.match(r"=SUM\(B3:B\d+\)$", b.replace(" ", "")):
                ws.cell(r, 2).value = f"=SUM(B3:B{r - 1})"
            break


def _create_month_sheet(wb, sheet_name: str):
    """Create a fresh single-month sheet (no prior-month carry-over): grand-total
    header + column labels only. Day columns, 负责人 groups, debtor rows and their
    prefix-SUMIF subtotals are all added on demand by the merge. Assumes a
    standalone current-month sheet (other sheets are irrelevant)."""
    ws = wb.create_sheet(title=sheet_name)
    ws.cell(1, 3).value = f"=SUM(C3:C{_EXCEL_MAX_ROW})"  # 余额合计
    ws.cell(1, 4).value = f"=SUM(D3:D{_EXCEL_MAX_ROW})"  # 当月发生额合计
    ws.cell(2, 2).value = "欠款人"
    ws.cell(2, 3).value = "欠款余额"
    ws.cell(2, 4).value = "当月发生额"
    ws.cell(2, 5).value = "上月余额"
    # colors from the example: C1 yellow / D1 pink totals, bold row-2 labels
    pr = _proto()
    _copy_style(pr.cell(1, 3), ws.cell(1, 3))
    _copy_style(pr.cell(1, 4), ws.cell(1, 4))
    for c in range(2, 6):
        _copy_style(pr.cell(2, c), ws.cell(2, c))
    return ws


# --------------------------------------------------------------------------- #
# hidden import log
# --------------------------------------------------------------------------- #
_LOG_HEADERS = ["欠单号", "日期", "欠款人", "初始金额", "月份sheet", "状态", "备注"]


def _ensure_log_sheet(wb):
    # migrate a legacy-named log sheet so past dedup history is preserved
    if LOG_SHEET not in wb.sheetnames:
        for n in _LEGACY_LOG_SHEETS:
            if n in wb.sheetnames:
                wb[n].title = LOG_SHEET
                break
    if LOG_SHEET in wb.sheetnames:
        ws = wb[LOG_SHEET]
    else:
        ws = wb.create_sheet(LOG_SHEET)
        ws.append(_LOG_HEADERS)
    ws.sheet_state = "hidden"
    return ws


def _log_sheet_name(wb) -> Optional[str]:
    """Current log sheet name, honoring the legacy name for un-migrated files."""
    if LOG_SHEET in wb.sheetnames:
        return LOG_SHEET
    for n in _LEGACY_LOG_SHEETS:
        if n in wb.sheetnames:
            return n
    return None


def _existing_iou_numbers(wb) -> set:
    name = _log_sheet_name(wb)
    if not name:
        return set()
    ws = wb[name]
    seen = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None:
            seen.add(str(v).strip())
    return seen


# --------------------------------------------------------------------------- #
# merge
# --------------------------------------------------------------------------- #
def merge(ae_bytes: bytes, qff_files: List[bytes]) -> Tuple[bytes, Dict[str, Any]]:
    """Merge QFF files into the AE ledger. Returns (updated_xlsx_bytes, report)."""
    wb = load_xlsx(ae_bytes)
    seen = _existing_iou_numbers(wb)

    # gather all IOUs across the uploaded QFF files
    ious: List[Dict[str, Any]] = []
    for data in qff_files:
        ious.extend(read_qff(data))

    report: Dict[str, Any] = {
        "total": len(ious),
        "added": 0,
        "aggregated": 0,
        "new_debtors": 0,
        "duplicates": 0,
        "new_sheets": 0,        # month sheets auto-created
        "new_sheet_list": [],   # ["202607", ...]
        "new_days": 0,          # day columns auto-created in month sheets
        "new_day_list": [],     # ["202607-260713", ...]
        "new_groups": 0,        # 负责人 groups auto-created
        "new_group_names": [],  # ["202607: WX", ...]
        "no_id": 0,             # rows imported without a 欠单号 (cannot be deduped)
        "skipped": [],          # {iou_no, reason}
        "new_debtor_names": [],
        "log_rows": [],         # rows to append to the hidden log
    }

    # Placement is driven by the 负责人 code in QFF column E: each IOU merges into
    # its own owner's group (QFF / WW / LYC / …) in the target month sheet.
    #
    # plan[sheet] = {
    #   "day_cols": {yymmdd: col},
    #   "owner_rows": {owner: {normname: row}},          # existing debtor rows per group
    #   "existing": [(row, col, amount, iou)],
    #   "new": {owner: {normname: {"name","cells","ious"}}},
    # }
    plan: Dict[str, Dict[str, Any]] = {}

    for iou in ious:
        no = iou["iou_no"]
        if no:  # has a 欠单号 -> dedup as before
            if no in seen:
                report["duplicates"] += 1
                report["skipped"].append({"iou_no": no, "reason": "duplicate"})
                continue
            seen.add(no)
        else:   # no 欠单号 -> always import, but cannot be tracked for dedup
            report["no_id"] += 1

        yymmdd = iou["date"]
        if not re.fullmatch(r"\d{6}", yymmdd or ""):
            report["skipped"].append({"iou_no": no, "reason": f"bad date {iou['date']!r}"})
            continue
        sheet_name = _date_to_ym_sheet(yymmdd)
        if sheet_name not in wb.sheetnames:
            _create_month_sheet(wb, sheet_name)
            report["new_sheets"] += 1
            report["new_sheet_list"].append(sheet_name)

        ws = wb[sheet_name]
        if sheet_name not in plan:
            plan[sheet_name] = {
                "day_cols": _build_day_columns(ws),
                "owner_rows": {},
                "existing": [],
                "new": {},
            }
        st = plan[sheet_name]

        col = st["day_cols"].get(yymmdd)
        if not col:
            # Auto-create the day column (safe: fills a pre-provisioned slot).
            col = _ensure_day_column(ws, yymmdd, st["day_cols"])
            if col:
                report["new_days"] += 1
                report["new_day_list"].append(f"{sheet_name}-{yymmdd}")
            else:
                report["skipped"].append({"iou_no": no, "reason": f"bad date for day column {yymmdd}"})
                continue

        owner = iou["owner"] or "QFF"
        if owner not in st["owner_rows"]:
            st["owner_rows"][owner] = _build_debtor_rows(ws, owner)
        # Missing groups are no longer skipped — they get created (see apply phase).

        amount = iou["initial"] or 0
        nk = _norm(iou["debtor"])
        row = st["owner_rows"][owner].get(nk)
        if row:
            st["existing"].append((row, col, amount, iou))
        else:
            grp = st["new"].setdefault(owner, {})
            nd = grp.setdefault(nk, {"name": iou["debtor"], "cells": {}, "ious": []})
            nd["cells"][col] = nd["cells"].get(col, 0) + amount
            nd["ious"].append(iou)

    # apply per sheet
    created = set(report["new_sheet_list"])  # sheets built fresh -> style from template
    for sheet_name, st in plan.items():
        ws = wb[sheet_name]
        fresh = sheet_name in created
        max_col = ws.max_column

        # 1) existing debtors: add amount into the day cell (aggregate).
        #    Done before any row insertion; the values move with their cells.
        for row, col, amount, iou in st["existing"]:
            cur = ws.cell(row, col).value
            base = cur if isinstance(cur, (int, float)) else 0
            ws.cell(row, col).value = base + amount
            report["added"] += 1
            if base:
                report["aggregated"] += 1
            report["log_rows"].append([iou["iou_no"], iou["date"], iou["debtor"],
                                       amount, sheet_name, "added", iou.get("remark")])

        # 2) new debtors: insert rows into each owner's group (creating the group
        #    if it doesn't exist yet). Process bottom-up (largest insertion row
        #    first) so earlier insertions don't shift the still-original points above.
        batches = []
        for owner, grp in st["new"].items():
            last = _last_group_row(ws, owner)
            if last is not None:                       # existing group -> append after it
                batches.append((last + 1, owner, list(grp.values()), False))
            else:                                      # new group -> just above 当月合计, with its own subtotal
                batches.append((_grand_total_row(ws), owner, list(grp.values()), True))
        batches.sort(key=lambda b: b[0], reverse=True)

        for insert_at, owner, new_list, is_new_group in batches:
            tpl = next((r for r in _build_debtor_rows(ws, owner).values() if r < insert_at), 3)
            n_rows = len(new_list) + (1 if is_new_group else 0)
            _insert_rows_with_remap(ws, insert_at, n_rows)
            for i, nd in enumerate(new_list):
                r = insert_at + i
                if fresh:
                    _style_debtor_row(ws, r, max_col)
                else:
                    _copy_row_style(ws, tpl, r, max_col)
                ws.cell(r, 1).value = owner
                ws.cell(r, 2).value = nd["name"]
                ws.cell(r, 3).value = f"=SUM(E{r}:ZC{r})"   # 余额
                ws.cell(r, 4).value = f"=SUM(F{r}:ZE{r})"   # 当月发生额
                for col, amt in nd["cells"].items():
                    ws.cell(r, col).value = amt
                report["new_debtors"] += 1
                report["new_debtor_names"].append(f"{owner}: {nd['name']}")
                for iou in nd["ious"]:
                    report["added"] += 1
                    report["log_rows"].append([iou["iou_no"], iou["date"], iou["debtor"],
                                               iou["initial"] or 0, sheet_name,
                                               "new-row", iou.get("remark")])
            if is_new_group:  # contiguous subtotal row for the brand-new group's block
                sr = insert_at + len(new_list)
                _style_subtotal_row(ws, sr)  # subtotal look (no debtor fill), any sheet
                first_r, last_r = insert_at, sr - 1
                ws.cell(sr, 1).value = f"=SUM(C{first_r}:C{last_r})"
                ws.cell(sr, 2).value = f"=SUM(D{first_r}:D{last_r})"
                report["new_groups"] += 1
                report["new_group_names"].append(f"{sheet_name}: {owner}")

        # 3) open-ended grand total + re-anchored day totals
        _open_grand_total(ws)

    # append to the hidden import log
    log = _ensure_log_sheet(wb)
    for row in report["log_rows"]:
        if row and str(row[0] or "").strip():  # only rows with a 欠单号 are dedup-tracked
            log.append(row)

    # Force Excel to recompute every formula on open — openpyxl writes no cached
    # values, so without this the formula cells would show blank until edited.
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:  # noqa: BLE001 — best-effort; older openpyxl
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # keep the report JSON-friendly
    report.pop("log_rows", None)
    return buf.getvalue(), report
