"""
One-off build script: trim `docs/2026.06 YUHANG.xlsx` into a clean, self-contained
template for the GJP 发票导出 tool.

Keeps:
  * the 电子发票 data sheet — rows 1-7 (headers/notes/styling) + row 8 as a styled
    formula prototype, base-column values cleared, data rows 9+ deleted
  * the GL sheet (Flight/Refund/... -> G/L account), needed by the X column VLOOKUP

Changes:
  * X8 VLOOKUP repointed from the broken external [3]GL link to the internal GL sheet
  * AD8 (English Name VLOOKUP) cleared — its external link is broken and would
    otherwise require committing employee PII

Output: backend/src/assets/gjp_invoice_template.xlsx

Run:  backend/venv/Scripts/python.exe backend/build_gjp_template.py
"""
import os
from copy import copy
import openpyxl
from openpyxl.worksheet.views import Selection

HERE = os.path.dirname(__file__)
SRC = os.path.join(HERE, "..", "docs", "2026.06 YUHANG.xlsx")
OUT = os.path.join(HERE, "src", "assets", "gjp_invoice_template.xlsx")

DATA_SHEET = "电子发票"
KEEP = {DATA_SHEET, "GL"}

# base columns (by letter) that must be blanked on the prototype row
BASE_COLS = ["C", "D", "E", "F", "G", "H", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]


def main() -> None:
    wb = openpyxl.load_workbook(SRC)

    # drop every sheet except the ones we keep
    for name in list(wb.sheetnames):
        if name not in KEEP:
            del wb[name]

    ws = wb[DATA_SHEET]

    # delete all data rows below the prototype (row 8)
    if ws.max_row > 8:
        ws.delete_rows(9, ws.max_row - 8)

    # repoint the G/L VLOOKUP to the internal GL sheet; clear the EN-name lookup
    ws["X8"].value = "=VLOOKUP(C8,GL!A:B,2,FALSE)"
    ws["AD8"].value = None

    # widen the summary sums (were capped at row 398) so a larger month still totals
    ws["F1"].value = "=SUM(Y8:Y2000)"
    ws["F2"].value = "=SUM(N8:N2000)"

    # hide the GL lookup sheet — its VLOOKUP still works while hidden, so the
    # exported file looks like a single 电子发票 tab
    wb["GL"].sheet_state = "hidden"

    # add a 电子客票号码 header in the (unused) U column, styled like the others
    ws["U7"].value = "电子客票号码"
    if ws["T7"].has_style:
        ws["U7"]._style = copy(ws["T7"]._style)

    # open at the top-left, not scrolled to where the old data used to end
    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
    wb.active = wb.sheetnames.index(DATA_SHEET)

    # blank the base-column values on the prototype row (keep styling + formulas)
    for col in BASE_COLS:
        ws[f"{col}8"].value = None
    ws["A8"].value = None  # 序号

    # drop external-link definitions so the file is self-contained
    try:
        wb._external_links = []
    except Exception:
        pass

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("wrote", os.path.abspath(OUT))
    print("sheets:", wb.sheetnames, "max_row:", ws.max_row)


if __name__ == "__main__":
    main()
