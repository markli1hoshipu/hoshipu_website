import json
import openpyxl
import re

def get_data(filename = 'client_atg.txt'):  
    # 载入数据
    with open(filename, 'r', encoding='utf8') as f:
        c = f.read()
        if c == '':
            cur_dict = {}
        else:
            cur_dict = json.loads(c)
        # 使用json格式存储的客户信息的字典,以两个公司为例子
        # 字典格式为：{'companies':[GROB],
        #             'iddata': {
        #                   'individuals':{'name':[[id,phone,engname],[id,phone,engname]]}, 处理重名, 以及便于后续通过engname索引到individuals
        #                   'GROB': {} 
        #                   }
        #              'passdata':{
        #                   'individuals':{'name':[[passport,phone,contact],[passport,phone,contact]]}, 其中phone和前面的一样
        #                   'GROB': {} 
        #                   }
        #             'settings':{
        #                   'uisettings':{}
        #                   'usersettings':{'phone':str,'remark':str, 'default_output':str, 'default_file':str, 'domestic_airlines' = list[str]}
        #                   'companydata':{
        #                       'GROB':{'phone':str, 'vip_code':{'CZ':str},'discount_code':{}}
        #                   }
        #                }
        #          }
    return cur_dict

# 不要改变大小写，有个别关键词要保持
# 其中航司名统一为大写
# 其中公司名考虑到有中文让用户自己区分，其他同理不变
def clean_excel_str(input):
    return str(input).strip()

def update_data(filename='client_atg.txt', dataname='client_atg_data.xlsx'):
    try:
        # 从excel读取数据更新文件，返回True,False,错误种类
        data = {'companies': [], 'iddata': {}, 'passdata': {}, 'settings': {}}
        book = openpyxl.load_workbook(dataname)
        sheetnames = book.sheetnames

        # 设置页
        sheet_settings = book[sheetnames[-1]]
        # ui设置
        data['settings']['uisettings'] = {}  # 暂时不读取 后续再说
        # 用户信息设置
        data['settings']['usersettings'] = {}
        data['settings']['usersettings']['phone'] = clean_excel_str(sheet_settings.cell(row=9, column=2).value)
        data['settings']['usersettings']['remark'] = clean_excel_str(sheet_settings.cell(row=10, column=2).value)
        data['settings']['usersettings']['default_output'] = clean_excel_str(sheet_settings.cell(row=11, column=2).value)
        data['settings']['usersettings']['default_file'] = clean_excel_str(sheet_settings.cell(row=13, column=2).value)
        data['settings']['usersettings']['domestic_airlines'] = [airline.strip().upper() \
                                                                 for airline in clean_excel_str(sheet_settings.cell(row=15, column=2).value).strip().replace('，',',').split(',')]
        # 公司设置
        data['settings']['companydata'] = {}
        row = 19  # 从17行开始有公司信息
        while clean_excel_str(sheet_settings.cell(row, 1).value) != 'END':
            cname = clean_excel_str(sheet_settings.cell(row, 1).value)
            data['settings']['companydata'][cname] = {}
            data['settings']['companydata'][cname]['phone'] = clean_excel_str(sheet_settings.cell(row, 2).value)
            data['settings']['companydata'][cname]['vip_code'] = {}
            for col in range(3, 6):
                airlines = clean_excel_str(sheet_settings.cell(18, col).value).split(',')
                for airline in airlines:
                    airline = airline.upper()
                    if clean_excel_str(sheet_settings.cell(row, col).value) != 'None':
                        data['settings']['companydata'][cname]['vip_code'][airline] = clean_excel_str(sheet_settings.cell(row, col).value)
            data['settings']['companydata'][cname]['discount_code'] = {}
            for col in range(6, 11):
                airlines = clean_excel_str(sheet_settings.cell(18, col).value).split(',')
                for airline in airlines:
                    airline = airline.upper()
                    if clean_excel_str(sheet_settings.cell(row, col).value) != 'None':
                        data['settings']['companydata'][cname]['discount_code'][airline] = clean_excel_str(sheet_settings.cell(row, col).value)
            row += 1

        # 散客页
        sheet_individual = book[sheetnames[0]]
        names = [cell.value for cell in sheet_individual['A'][2:]]
        ids = [cell.value for cell in sheet_individual['B'][2:]]
        phones = [cell.value for cell in sheet_individual['C'][2:]]
        engnames = [cell.value for cell in sheet_individual['D'][2:]]
        passports = [cell.value for cell in sheet_individual['E'][2:]]
        iphones = [cell.value for cell in sheet_individual['F'][2:]]
        iemails = [cell.value for cell in sheet_individual['G'][2:]]
        langugaes = [cell.value for cell in sheet_individual['H'][2:]]

        data['iddata']['individuals'] = {}
        data['passdata']['individuals'] = {}
        for i in range(len(names)):
            if clean_excel_str(names[i]) == 'END':
                break
            name = clean_excel_str(names[i])
            id = clean_excel_str(ids[i])
            phone = clean_excel_str(phones[i])
            engname = clean_excel_str(engnames[i])
            passport = clean_excel_str(passports[i])
            iphone = clean_excel_str(iphones[i])
            iemail = clean_excel_str(iemails[i])
            langugae = clean_excel_str(langugaes[i])
            if langugae == 'None': language = 'en'

            if name != 'None':
                if name not in data['iddata']['individuals']:
                    data['iddata']['individuals'][name] = []
                data['iddata']['individuals'][name].append([id, phone,engname])
            if engname != 'None':
                if engname not in data['passdata']['individuals']:
                    data['passdata']['individuals'][engname] = []
                data['passdata']['individuals'][engname].append([passport,phone,iphone,iemail,langugae])
        
        # 公司页
        for sheetname in sheetnames[1:-1]:
            sheet_company = book[sheetname]
            sheetname = clean_excel_str(sheetname)
            data['companies'].append(sheetname)
            names = [cell.value for cell in sheet_company['A'][2:]]
            ids = [cell.value for cell in sheet_company['B'][2:]]
            phones = [cell.value for cell in sheet_company['C'][2:]]
            engnames = [cell.value for cell in sheet_company['D'][2:]]
            passports = [cell.value for cell in sheet_company['E'][2:]]
            iphones = [cell.value for cell in sheet_individual['F'][2:]]
            iemails = [cell.value for cell in sheet_individual['G'][2:]]
            langugaes = [cell.value for cell in sheet_individual['H'][2:]]

            data['iddata'][sheetname] = {}
            data['passdata'][sheetname] = {}
            for i in range(len(names)):
                if clean_excel_str(names[i]) == 'END':
                    break
                name = clean_excel_str(names[i])
                id = clean_excel_str(ids[i])
                phone = clean_excel_str(phones[i])
                engname = clean_excel_str(engnames[i])
                passport = clean_excel_str(passports[i])
                iphone = clean_excel_str(iphones[i])
                iemail = clean_excel_str(iemails[i])
                langugae = clean_excel_str(langugaes[i])
                if langugae == 'None': language = 'en'

                if name != 'None':
                    if name not in data['iddata'][sheetname]:
                        data['iddata'][sheetname][name] = []
                    data['iddata'][sheetname][name].append([id, phone,engname])
                if engname != 'None':
                    if engname not in data['passdata'][sheetname]:
                        data['passdata'][sheetname][engname] = []
                    data['passdata'][sheetname][engname].append([passport,phone,iphone,iemail,langugae])
        with open(filename, 'w', encoding='utf8') as f:
            json.dump(data, f)
        book.close()
        return True
    except Exception as e:
        # 如果操作发生错误，返回错误信息
        return str(e)

def update_data(filename='client_atg.txt', dataname='client_atg_data.xlsx'):
    # 从excel读取数据更新文件，返回True,False,错误种类
    data = {'companies': [], 'iddata': {}, 'passdata': {}, 'settings': {}}
    book = openpyxl.load_workbook(dataname)
    sheetnames = book.sheetnames

    # 设置页
    sheet_settings = book[sheetnames[-1]]
    # ui设置
    data['settings']['uisettings'] = {}  # 暂时不读取 后续再说
    # 用户信息设置
    data['settings']['usersettings'] = {}
    data['settings']['usersettings']['phone'] = clean_excel_str(sheet_settings.cell(row=9, column=2).value)
    data['settings']['usersettings']['remark'] = clean_excel_str(sheet_settings.cell(row=10, column=2).value)
    data['settings']['usersettings']['default_output'] = clean_excel_str(sheet_settings.cell(row=11, column=2).value)
    data['settings']['usersettings']['default_file'] = clean_excel_str(sheet_settings.cell(row=13, column=2).value)
    data['settings']['usersettings']['domestic_airlines'] = [airline.strip().upper() \
                                                                for airline in clean_excel_str(sheet_settings.cell(row=15, column=2).value).strip().replace('，',',').split(',')]
    # 公司设置
    data['settings']['companydata'] = {}
    row = 19  # 从17行开始有公司信息
    while clean_excel_str(sheet_settings.cell(row, 1).value) != 'END':
        cname = clean_excel_str(sheet_settings.cell(row, 1).value)
        data['settings']['companydata'][cname] = {}
        data['settings']['companydata'][cname]['phone'] = clean_excel_str(sheet_settings.cell(row, 2).value)
        data['settings']['companydata'][cname]['vip_code'] = {}
        for col in range(3, 6):
            airlines = clean_excel_str(sheet_settings.cell(18, col).value).split(',')
            for airline in airlines:
                airline = airline.upper()
                if clean_excel_str(sheet_settings.cell(row, col).value) != 'None':
                    data['settings']['companydata'][cname]['vip_code'][airline] = clean_excel_str(sheet_settings.cell(row, col).value)
        data['settings']['companydata'][cname]['discount_code'] = {}
        for col in range(6, 11):
            airlines = clean_excel_str(sheet_settings.cell(18, col).value).split(',')
            for airline in airlines:
                airline = airline.upper()
                if clean_excel_str(sheet_settings.cell(row, col).value) != 'None':
                    data['settings']['companydata'][cname]['discount_code'][airline] = clean_excel_str(sheet_settings.cell(row, col).value)
        row += 1

    # 散客页
    sheet_individual = book[sheetnames[0]]
    names = [cell.value for cell in sheet_individual['A'][2:]]
    ids = [cell.value for cell in sheet_individual['B'][2:]]
    phones = [cell.value for cell in sheet_individual['C'][2:]]
    engnames = [cell.value for cell in sheet_individual['D'][2:]]
    passports = [cell.value for cell in sheet_individual['E'][2:]]
    iphones = [cell.value for cell in sheet_individual['F'][2:]]
    iemails = [cell.value for cell in sheet_individual['G'][2:]]
    langugaes = [cell.value for cell in sheet_individual['H'][2:]]

    data['iddata']['individuals'] = {}
    data['passdata']['individuals'] = {}
    for i in range(len(names)):
        if clean_excel_str(names[i]) == 'END':
            break
        name = clean_excel_str(names[i])
        id = clean_excel_str(ids[i])
        phone = clean_excel_str(phones[i])
        engname = clean_excel_str(engnames[i])
        passport = clean_excel_str(passports[i])
        iphone = clean_excel_str(iphones[i])
        iemail = clean_excel_str(iemails[i])
        langugae = clean_excel_str(langugaes[i])
        if langugae == 'None': 
            language = 'en'

        if name != 'None':
            if name not in data['iddata']['individuals']:
                data['iddata']['individuals'][name] = []
            data['iddata']['individuals'][name].append([id, phone,engname])
        if engname != 'None':
            if engname not in data['passdata']['individuals']:
                data['passdata']['individuals'][engname] = []
            data['passdata']['individuals'][engname].append([passport,phone,iphone,iemail,langugae])
    # 公司页
    for sheetname in sheetnames[1:-1]:
        sheet_company = book[sheetname]
        sheetname = clean_excel_str(sheetname)
        data['companies'].append(sheetname)
        names = [cell.value for cell in sheet_company['A'][2:]]
        ids = [cell.value for cell in sheet_company['B'][2:]]
        phones = [cell.value for cell in sheet_company['C'][2:]]
        engnames = [cell.value for cell in sheet_company['D'][2:]]
        passports = [cell.value for cell in sheet_company['E'][2:]]
        iphones = [cell.value for cell in sheet_company['F'][2:]]
        iemails = [cell.value for cell in sheet_company['G'][2:]]
        langugaes = [cell.value for cell in sheet_company['H'][2:]]
        data['iddata'][sheetname] = {}
        data['passdata'][sheetname] = {}
        for i in range(len(names)):
            if clean_excel_str(names[i]) == 'END':
                break
            name = clean_excel_str(names[i])
            id = clean_excel_str(ids[i])
            phone = clean_excel_str(phones[i])
            engname = clean_excel_str(engnames[i])
            passport = clean_excel_str(passports[i])
            iphone = clean_excel_str(iphones[i])
            iemail = clean_excel_str(iemails[i])
            langugae = clean_excel_str(langugaes[i])
            if langugae == 'None': 
                language = 'en'

            if name != 'None':
                if name not in data['iddata'][sheetname]:
                    data['iddata'][sheetname][name] = []
                data['iddata'][sheetname][name].append([id, phone,engname])
            if engname != 'None':
                if engname not in data['passdata'][sheetname]:
                    data['passdata'][sheetname][engname] = []
                data['passdata'][sheetname][engname].append([passport,phone,iphone,iemail,langugae])
    with open(filename, 'w', encoding='utf8') as f:
        json.dump(data, f)
    book.close()
    return True

   
# 在生成部分 考虑到要让用户选择重名问题 把生成函数分为两个：一主一次

def generate_output_headtail(user_phone,company,airline,id_type,flighttype,remark,user_input,data_dict):
    lines = user_input.splitlines()
    names = lines[2:]
    input_names = ''
    for n in names:
        input_names += '1'+n
    head = ''; tail = ''
    head += 'sd' + lines[0][-3:] + '\n'
    head += f'nm{input_names}\n'
    head += f'ct {remark}\n'
    if airline in data_dict['settings']['usersettings']['domestic_airlines']:
        head += f'OSI {airline} CTCT{user_phone}\n'


    # flight type 是 1 为国内，2为国际
    if company in data_dict['companies']:
        if flighttype == 1 and data_dict['settings']['companydata'][company]['phone'] != 'None': 
            tail += f'rmk mp {data_dict['settings']['companydata'][company]['phone']}\n'
        if airline in data_dict['settings']['companydata'][company]['vip_code']:
            tail += f'RMK IC {airline}/{data_dict['settings']['companydata'][company]['vip_code'][airline]}\n'
    tail += f'tktl1800/./dlc159\n'
    tail += '@ki\n'

    tail += '\n以下为查询价格指令:\n\n'
    if company in data_dict['companies']:
        if airline in data_dict['settings']['companydata'][company]['discount_code']:
            tail += f'PAT:A{data_dict['settings']['companydata'][company]['discount_code'][airline]}\n\n'
        else:
            tail +='PAT:A\n\n'
    else:
        tail +='PAT:A\n\n'
    return head,tail

def generate_output_name(name,rank,company,airline,id_type,flighttype,data_dict):
    # 返回最后要显示的内容，或者报错信息
    # 均不含有/P几
    # 其中id_type为1则是身份证，为2则是护照
    # flight_type为1则是国内，为2 是国际，由于设置是不可能出现国际+身份证不需要考虑
    if company not in data_dict['companies']:
        company = 'individuals'
    fre = []
    # 这里情况非常繁杂, 为了保持清晰就分情况
    # 首先大体上一定会有ctcm和ssr信息项
    # 然后在国内的航班会有rmk项
    try:
        if id_type == 1: # 优先身份证，默认是国内航司+国内航班, 无视判定
            pcustomers = []
            if name in data_dict['iddata'][company]:
                pcustomers += data_dict['iddata'][company][name]
            if len(pcustomers) == 0: #至少没有找到直接由身份证的，不考虑比如MARKLI同时有身份证，而李智远有英文名MARKLI，即不支持英文名反向索引身份证
                if name in data_dict['passdata'][company]:
                    pcustomers += data_dict['passdata'][company][name]
                if len(pcustomers) == 0:
                    return 0, fre #这是真的啥也没有
                else:
                    # 这个情况比较奇怪，所有单独拿出来写
                    # 首先返回的必须是0,? 来表示首位索引没有
                    # 其次输入略有不同
                    print(pcustomers)
                    for customer in pcustomers: #customer 为[passport,phone,iphone,iemail,langugae]
                        re = ''
                        re += f'OSI {airline} CTCM{customer[1]}/P{rank}\n' #这个预留出来方便修改
                        re += f'SSR DOCS {airline} HK1 P/{customer[0]}/P{rank}\n'
                        if flighttype == 1 and customer [1] != 'None': #这个如果没有就确实没有吧
                            re += f'rmk mp {customer[1]}/P{rank}\n'
                        fre.append(re)
                    return 0, fre
            else:  #有直接身份证的就身份证了
                for customer in pcustomers: #customer 为[id,phone,engname]
                    re = ''
                    re += f'OSI {airline} CTCM{customer[1]}/P{rank}\n'
                    re += f'SSR FOID {airline} HK/NI{customer[0]}/P{rank}\n'
                    re += f'rmk mp {customer[1]}/P{rank}\n'
                    fre.append(re)
                if len(fre) > 1:
                    return 2,fre
                return 1,fre

        elif id_type == 2: # 优先护照
            pcustomers = []
            if name in data_dict['passdata'][company]:
                pcustomers += data_dict['passdata'][company][name]
            if name in data_dict['iddata'][company]:
                for cus in data_dict['iddata'][company][name]:
                    if cus[-1] != 'None' and cus[-1] in data_dict['passdata'][company]: # cus[-1] 为engname
                        pcustomers.append(data_dict['passdata'][company][cus[-1]])
            if len(pcustomers) == 0:
                return 0, fre
            for customer in pcustomers: #customer 为[passport,phone,iphone,iemail,langugae]
                re = ''
                if airline in data_dict['settings']['usersettings']['domestic_airlines']: #国内航司, 只ctcm国内手机号
                    re += f'OSI {airline} CTCM{customer[1]}/P{rank}\n'
                else: #境外航司
                    if customer[2] != 'None': #国际手机,-1为语言
                        re += f'SSR CTCM {airline} HK1 {customer[2]}/{customer[-1]}/P{rank}\n' 
                    if customer[3] != 'None': #国际邮箱，-1为语言
                        re += f'SSR CTCE {airline} HK1 {customer[3]}/{customer[-1]}/P{rank}\n'
                    if customer[2] == 'None' and customer[3] == 'None': #实际上应该报错 但是为了方便允许自己写吧
                        if customer [1] != 'None': #默认中国手机号
                            re += f'SSR CTCM {airline} HK1 86{customer[1]}/{customer[-1]}/P{rank}\n' 
                        else:
                            re += f'SSR CTCM {airline} HK1 没有联系方式自行填写手机/{customer[-1]}/P{rank}\n' 
                            re += f'SSR CTCE {airline} HK1 没有联系方式自行填写邮箱/{customer[-1]}/P{rank}\n'
                re += f'SSR DOCS {airline} HK1 P/{customer[0]}/P{rank}\n'
                if flighttype == 1 and customer [1] != 'None':
                    re += f'rmk mp {customer[1]}/P{rank}\n'
                fre.append(re)
            if len(fre) > 1:
                return 2,fre
            return 1,fre
        
        return 4, ['选择输出种类有问题！']
    except:
        return 4, ['选择输出种类有问题！']

def format_lines(text):
    # 取消所有换行符
    text = text.replace('\n', '').replace('/t','')
    while '  ' in text:
        text = text.replace('  ',' ')
    
    # 在每个数字前添加换行符
    numbers = [str(i) + '.' for i in range(1, 100)]
    for number in numbers:
        if number == '26.':
            continue
        text = text.replace(number, '\n' + number,1)

    return text.strip()

def rt_to_customers(rt_input,data_dict):
#     # rt输入的主要问题是如何判断人数，其实一个简单的办法是: rt人数不可能超过9人，所以一个一个找就行
#     # 而且因为无论如何都要新开行数，所以每个customers直接按照
#     # name, id, phone, engname, pass, iphone, iemail, language 就行

#     # rt出来的信息实在是太奇怪...
#     # 必须先规范好


    rt_input = format_lines(rt_input)
    if re.compile(r'OSI (..)',re.M).findall(rt_input) != []:
        airline = re.compile(r'OSI (..)',re.M).findall(rt_input)[0].strip()
    else:
        airline = re.compile(r'SSR DOCS (..)',re.M).findall(rt_input)[0].strip()
    ret = []
    count = 0
    for rank in range(1,10):
        name = ''; engname = ''; id = ''; passport = ''

        id = re.compile(rf'SSR FOID {airline} HK1 NI(.*?)/P{rank}',re.M).findall(rt_input)
        if id != []:
            id = id[0].strip()
        else:
            id = ''
        passport = re.compile(rf'SSR DOCS {airline} HK1 P/(.*?)/P{rank}',re.M).findall(rt_input)
        if passport != []: 
            passport = passport[0].strip()
        else:
            passport = ''
        if id == '' and passport == '':
            count = rank-1
            break
        phone = ''; iphone = ''; iemail = ''; language = ''

        if airline in airline in data_dict['settings']['usersettings']['domestic_airlines']:
            phone = re.compile(rf'OSI {airline} CTCM(.*)/P{rank}',re.M).findall(rt_input)[0]
            iphone = '86'+phone
        else:
            iphone = re.compile(rf'SSR CTCM {airline} HK1 (.*)/P{rank}',re.M).findall(rt_input)
            if iphone != []:
                iphone = iphone[0]
                if iphone[-3] == '/':
                    language = iphone[-2:]
                    iphone = iphone[:-3]
            else:
                iphone = ''
            iemail = re.compile(rf'SSR CTCE {airline} HK1 (.*)/P{rank}',re.M).findall(rt_input)
            if iemail != []:
                iemail = iemail[0]
                if iemail[-3] == '/':
                    language = iemail[-2:]
                    iemail = iemail[:-3]
            else:
                iemail = ''
        ret.append([name, id, phone, engname, passport, iphone, iemail, language])
        
    for c in range(0,count-1):
        name = re.compile(rf'{c+1}\.(.*?)\n',re.M).findall(rt_input)[0].strip()
        if ret[c][1] == '':
            ret[c][3] = name #英文
        else:
            ret[c][0] = name #中文
    name = re.compile(rf'{count}\.(.*?)\n',re.M).findall(rt_input)[0].strip()[:-6].strip() #排除记录号
    if ret[count-1][1] == '':
        ret[count-1][3] = name #英文
    else:
        ret[count-1][0] = name #中文
    return ret

def upload_new_data(rt_input,sheetname,data_dict,dataname='client_atg_data.xlsx'):
    # localdata 从excel在主程序处做更新就可以了
    try:
        customers = rt_to_customers(rt_input,data_dict)
        if len(customers) == 0:
            return '未识别到客户！'
        book = openpyxl.load_workbook(dataname)
        if sheetname not in book.sheetnames:
            return '没有这个sheet！'
        sheet = book[sheetname]
        row = 1
        while sheet.cell(row=row, column=1).value != 'END':
            row += 1
        
        for customer in customers:
            for col in range(8):
                sheet.cell(row=row, column=col+1).value = customer[col]
            row +=1 
        for col in range(8):
            sheet.cell(row=row, column=col+1).value = 'END'
        book.save(dataname)
        return True
    except Exception as e:
        # 如果操作发生错误，返回错误信息
        return str(e)
