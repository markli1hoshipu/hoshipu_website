import xlrd,time,pickle,openpyxl,copy
from openpyxl.styles import *

#每个business格式为：
#(name,client,month,[airout,airoutp],[airin,airinp],[aircus,aircusp],\
# [seaout,seaoutp],[seain,seainp],[seacus,seacusp],[fasttr,fasttrp],\
# [railout,railoutp],[railin,railinp],[domland,domlandp])

#general funcs
def nofr_sum(sh,column_n, row):
    r = 2; re = 0
    while r <= row:
        re += sh.cell(r,column_n).value
        r += 1
    return re

def fsum(lp):
    re = 0
    for i in lp:
        re+= i[0]
    return re
    
def ssum(lp):
    re = 0
    for i in lp:
        re+= i[1]
    return re

def bfloat(var):
    if var == '':
        return float(0)
    else:
        return float(var)

def get_translate(filename = 'translate.xlsx'):
    tr = {}
    book = xlrd.open_workbook(filename)
    sheet = book.sheet_by_index(0)
    keys = sheet.col_values(0)
    vals = sheet.col_values(1)
    for i in range(len(keys)):
        tr[keys[i]] = vals[i]
    return tr

#修改实际输出文本的样式
def fullborder(ws,l):
    #全部框线
    area = ws[f'A1:AI{l}']
    for i in area:
        for j in i:
            j.border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
            font = Font(name="等线",size=8,bold=False,italic=False,color="000000")
            j.font = font
    
    #十一个三组列
    areas = [ws[f'A1:C{l}'],ws[f'D1:F{l}'],ws[f'G1:I{l}'],ws[f'J1:L{l}'],ws[f'M1:O{l}'],\
    ws[f'P1:R{l}'],ws[f'S1:U{l}'],ws[f'V1:X{l}'],ws[f'Y1:AA{l}'],ws[f'AB1:AD{l}'],ws[f'AE1:AG{l}']]
    colors = ['ffffff','ffebcd','98f5ff','8470ff','7ccd7c','e6e6fa','eedc82','dda0dd','ee9572','ee6aa7','e0eee0']
    ct = 0
    for area in areas:
        for i in area:
            for j in i:
                pattern_fill = PatternFill(fill_type="solid",fgColor=colors[ct])
                j.fill = pattern_fill
        ct+=1

    #倒数第一列
    area = ws[f'AI1:AI{l}']
    for i in area:
        for j in i:
            pattern_fill = PatternFill(fill_type="solid",fgColor="cfcfcf")
            j.fill = pattern_fill

    #首行
    area = ws[f'A1:AI1']
    for i in area:
        for j in i:
            j.border = Border(right=Side(style='thick'),bottom=Side(style='thick'))
            font = Font(name="等线",size=8,bold=True,italic=False,color="FFFFFF")
            j.font = font
            pattern_fill = PatternFill(fill_type="solid",fgColor="000000")
            j.fill = pattern_fill

    return ws

def getsums(bb):
    firstsum = 0; secondsum = 0
    firstsum += bb[3][0]; secondsum += bb[3][0]*bb[3][1]
    firstsum += bb[4][0]; secondsum += bb[4][0]*bb[4][1]
    firstsum += bb[5][0]; secondsum += bb[5][0]*bb[5][1]

    firstsum += bb[6][0]; secondsum += bb[6][0]*bb[6][1]
    firstsum += bb[7][0]; secondsum += bb[7][0]*bb[7][1]
    firstsum += bb[8][0]; secondsum += bb[8][0]*bb[8][1]

    firstsum += bb[9][0]; secondsum += bb[9][0]*bb[9][1]

    firstsum += bb[10][0]; secondsum += bb[10][0]*bb[10][1]
    firstsum += bb[11][0]; secondsum += bb[11][0]*bb[11][1]
    firstsum += bb[12][0]; secondsum += bb[12][0]*bb[12][1]
    return firstsum,secondsum

#功能一：合并

#打开filename中sheetname一页, 返回一个list_of_business
def read_table(filename,sheetname = 'marklihoshipu'):

    # 锁定sheet
    book = xlrd.open_workbook(filename)
    
    if sheetname == 'marklihoshipu' or sheetname not in book.sheet_names():
        sheet = book.sheet_by_index(0)
    else:
        sheet = book.sheet_by_name(sheetname)

    re = []
    #读取business
    rowct = 0
    print()
    while sheet.cell(rowct+1, 0).value != 'END':
        rowct += 1
        re.append([ str(sheet.cell(rowct, 0).value).strip(),  #name
        str(sheet.cell(rowct, 1).value).strip(),              #client
        str(sheet.cell(rowct,  2).value).strip(),              #month
        [bfloat(sheet.cell(rowct,  3).value),                  #airout
        bfloat(sheet.cell( rowct,  4).value)],                  #airoutp
        [bfloat(sheet.cell( rowct,  6).value),                  
        bfloat(sheet.cell( rowct,  7).value)],                  
        [bfloat(sheet.cell( rowct,  9).value),                  
        bfloat(sheet.cell( rowct,  10).value)],                  
        [bfloat(sheet.cell( rowct,  12).value),                  
        bfloat(sheet.cell( rowct,  13).value)],     
        [bfloat(sheet.cell( rowct,  15).value),                  
        bfloat(sheet.cell( rowct,  16).value)],                  
        [bfloat(sheet.cell( rowct,18).value),                  
        bfloat(sheet.cell( rowct,  19).value)],                  
        [bfloat(sheet.cell(rowct,  21).value),                  
        bfloat(sheet.cell( rowct,  22).value)], 
        [bfloat(sheet.cell(rowct,  24).value),                  
        bfloat(sheet.cell( rowct,  25).value)],                  
        [bfloat(sheet.cell( rowct, 27).value),                  
        bfloat(sheet.cell( rowct,  28).value)],                  
        [bfloat(sheet.cell( rowct, 30).value),                  
        bfloat(sheet.cell(rowct, 31).value)],              
        ])
    return re

def read_table_zh(filename,sheetname = 'marklihoshipu',translate={}):

    # 锁定sheet
    book = xlrd.open_workbook(filename)
    
    if sheetname == 'marklihoshipu' or sheetname not in book.sheet_names():
        sheet = book.sheet_by_index(0)
    else:
        sheet = book.sheet_by_name(sheetname)

    re = []
    #读取business
    rowct = 0
    print()
    while sheet.cell(rowct+1, 0).value != 'END':
        rowct += 1
        client = str(sheet.cell(rowct, 1).value).strip()
        if client in translate:
            client = translate[client]
        re.append([ str(sheet.cell(rowct, 0).value).strip(),  #name
        client,              #client
        str(sheet.cell(rowct,  2).value).strip(),              #month
        [bfloat(sheet.cell(rowct,  3).value),                  #airout
        bfloat(sheet.cell( rowct,  4).value)],                  #airoutp
        [bfloat(sheet.cell( rowct,  6).value),                  
        bfloat(sheet.cell( rowct,  7).value)],                  
        [bfloat(sheet.cell( rowct,  9).value),                  
        bfloat(sheet.cell( rowct,  10).value)],                  
        [bfloat(sheet.cell( rowct,  12).value),                  
        bfloat(sheet.cell( rowct,  13).value)],     
        [bfloat(sheet.cell( rowct,  15).value),                  
        bfloat(sheet.cell( rowct,  16).value)],                  
        [bfloat(sheet.cell( rowct,18).value),                  
        bfloat(sheet.cell( rowct,  19).value)],                  
        [bfloat(sheet.cell(rowct,  21).value),                  
        bfloat(sheet.cell( rowct,  22).value)], 
        [bfloat(sheet.cell(rowct,  24).value),                  
        bfloat(sheet.cell( rowct,  25).value)],                  
        [bfloat(sheet.cell( rowct, 27).value),                  
        bfloat(sheet.cell( rowct,  28).value)],                  
        [bfloat(sheet.cell( rowct, 30).value),                  
        bfloat(sheet.cell(rowct, 31).value)],              
        ])
    return re


def write_table(filename,list_of_business):
    book = openpyxl.Workbook()
    sh = book.active
    sh.title = '汇总表格'

    #表头部分
    sh.cell(1,1).value = '销售' ; sh.cell(1,2).value = '客户' ; sh.cell(1,3).value = '月份'

    sh.cell(1,4).value = '空运出口';sh.cell(1,5).value = '计提比例';sh.cell(1,6).value = '计提利润'
    sh.cell(1,7).value = '空运进口'; sh.cell(1,8).value = '计提比例'; sh.cell(1,9).value = '计提利润'
    sh.cell(1,10).value = '空运关务'; sh.cell(1,11).value = '计提比例'; sh.cell(1,12).value = '计提利润'

    sh.cell(1,13).value = '海运出口';sh.cell(1,14).value = '计提比例';sh.cell(1,15).value = '计提利润'
    sh.cell(1,16).value = '海运进口'; sh.cell(1,17).value = '计提比例'; sh.cell(1,18).value = '计提利润'
    sh.cell(1,19).value = '海运关务'; sh.cell(1,20).value = '计提比例'; sh.cell(1,21).value = '计提利润'

    sh.cell(1,22).value = '快运' ; sh.cell(1,23).value = '计提比例' ; sh.cell(1,24).value = '计提利润'

    sh.cell(1,25).value = '铁路出口';sh.cell(1,26).value = '计提比例';sh.cell(1,27).value = '计提利润'
    sh.cell(1,28).value = '铁路进口'; sh.cell(1,29).value = '计提比例'; sh.cell(1,30).value = '计提利润'
    sh.cell(1,31).value = '国内陆运'; sh.cell(1,32).value = '计提比例'; sh.cell(1,33).value = '计提利润'
    
    sh.cell(1,34).value = '毛利润合计' ; sh.cell(1,35).value = '计提利润合计' ; 

    #内容部分
    row = 1
    firstsum = 0
    secondsum = 0
    for bb in list_of_business:
        row += 1
        sh.cell(row,1).value = bb[0] ; sh.cell(row,2).value = bb[1] ; sh.cell(row,3).value = bb[2]

        sh.cell(row,4).value = bb[3][0];sh.cell(row,5).value = bb[3][1];sh.cell(row,6).value = bb[3][0]*bb[3][1]; firstsum += bb[3][0]; secondsum += bb[3][0]*bb[3][1]
        sh.cell(row,7).value = bb[4][0]; sh.cell(row,8).value = bb[4][1]; sh.cell(row,9).value = bb[4][0]*bb[4][1]; firstsum += bb[4][0]; secondsum += bb[4][0]*bb[4][1]
        sh.cell(row,10).value = bb[5][0]; sh.cell(row,11).value = bb[5][1]; sh.cell(row,12).value = bb[5][0]*bb[5][1]; firstsum += bb[5][0]; secondsum += bb[5][0]*bb[5][1]

        sh.cell(row,13).value = bb[6][0];sh.cell(row,14).value = bb[6][1];sh.cell(row,15).value = bb[6][0]*bb[6][1]; firstsum += bb[6][0]; secondsum += bb[6][0]*bb[6][1]
        sh.cell(row,16).value = bb[7][0]; sh.cell(row,17).value = bb[7][1]; sh.cell(row,18).value = bb[7][0]*bb[7][1]; firstsum += bb[7][0]; secondsum += bb[7][0]*bb[7][1]
        sh.cell(row,19).value = bb[8][0]; sh.cell(row,20).value = bb[8][1]; sh.cell(row,21).value = bb[8][0]*bb[8][1]; firstsum += bb[8][0]; secondsum += bb[8][0]*bb[8][1]

        sh.cell(row,22).value = bb[9][0] ; sh.cell(row,23).value = bb[9][1] ; sh.cell(row,24).value = bb[9][0]*bb[9][1]; firstsum += bb[9][0]; secondsum += bb[9][0]*bb[9][1]

        sh.cell(row,25).value = bb[10][0];sh.cell(row,26).value = bb[10][1];sh.cell(row,27).value = bb[10][0]*bb[10][1]; firstsum += bb[10][0]; secondsum += bb[10][0]*bb[10][1]
        sh.cell(row,28).value = bb[11][0]; sh.cell(row,29).value = bb[11][1]; sh.cell(row,30).value = bb[11][0]*bb[11][1]; firstsum += bb[11][0]; secondsum += bb[11][0]*bb[11][1]
        sh.cell(row,31).value = bb[12][0]; sh.cell(row,32).value = bb[12][1]; sh.cell(row,33).value = bb[12][0]*bb[12][1]; firstsum += bb[12][0]; secondsum += bb[12][0]*bb[12][1]
        
        for i in range(5,33,3):
            sh.cell(row,i).number_format = '0%'

        sh.cell(row,34).value = firstsum ; sh.cell(row,35).value = secondsum ; 
        firstsum = 0
        secondsum = 0

    row += 1
    sh.cell(row,1).value = 'END'

    sh = fullborder(sh,row)
    book.save(f'{filename}.xlsx')

def write_table_from_filenames(destfilenname,filenames,sheetnames):
    list_of_business = []
    for i in range(min(len(filenames),len(sheetnames))):
        list_of_business += read_table(filenames[i],sheetname=sheetnames[i])
    write_table(destfilenname,list_of_business)
    return list_of_business

def write_table_from_filenames_zh(destfilenname,filenames,sheetnames,translate):
    list_of_business = []
    for i in range(min(len(filenames),len(sheetnames))):
        list_of_business += read_table_zh(filenames[i],sheetname=sheetnames[i],translate=translate)
    write_table(destfilenname,list_of_business)
    return list_of_business


def write_table0(dest, sheetname, list_of_business):
    
    wb = openpyxl.load_workbook(dest+'.xlsx')
    wb.create_sheet(sheetname)
    sh = wb[sheetname]
    sh.column_dimensions['A'].width = 12.0
    sh.column_dimensions['B'].width = 12.0
    sh.column_dimensions['C'].width = 12.0

    info = {}
    for business in list_of_business:
        if business[0] not in info:
            info[business[0]] = [0,0]
        info[business[0]][0] += getsums(business)[0]
        info[business[0]][1] += getsums(business)[1]
    
    #表头部分
    sh.cell(1,1).value = '销售' ; sh.cell(1,2).value = '毛利润合计' ; sh.cell(1,3).value = '计提利润合计'

    border = Border(right=Side(style='thick'),bottom=Side(style='thick'))
    font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
    pattern_fill = PatternFill(fill_type="solid",fgColor="c6e2ff")
    sh.cell(1,1).border = border;sh.cell(1,2).border = border;sh.cell(1,3).border = border
    sh.cell(1,1).font = font;sh.cell(1,2).font = font;sh.cell(1,3).font = font
    sh.cell(1,1).fill = pattern_fill;sh.cell(1,2).fill = pattern_fill;sh.cell(1,3).fill = pattern_fill

    #内容部分
    row = 1
    for name,valpair in info.items():
        row += 1
        sh.cell(row,1).value = name; sh.cell(row,2).value = valpair[0] ; sh.cell(row,3).value = valpair[1]
        border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
        font = Font(name="等线",size=8,bold=False,italic=False,color="000000")
        sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
        sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font

    row+=1
    sh.cell(row,1).value = '合计'

    sh.cell(row,2).value = nofr_sum(sh,2,row-1)
    sh.cell(row,3).value = nofr_sum(sh,3,row-1)

    border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
    font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
    pattern_fill = PatternFill(fill_type="solid",fgColor="9f79ee")
    sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
    sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font
    sh.cell(row,1).fill = pattern_fill;sh.cell(row,2).fill = pattern_fill;sh.cell(row,3).fill = pattern_fill



    wb.save(dest+'.xlsx')
    print('zhuijia')

def write_table1(dest, sheetname, list_of_business):
    
    wb = openpyxl.load_workbook(dest+'.xlsx')
    wb.create_sheet(sheetname)
    sh = wb[sheetname]
    sh.column_dimensions['A'].width = 12.0
    sh.column_dimensions['B'].width = 12.0
    sh.column_dimensions['C'].width = 12.0
    info = {}
    for business in list_of_business:
        if business[0] not in info:
            info[business[0]] = {}
        if business[2] not in info[business[0]]:
            info[business[0]][business[2]] = [0,0]
        info[business[0]][business[2]][0] += getsums(business)[0]
        info[business[0]][business[2]][1] += getsums(business)[1]
    
    #表头部分
    sh.cell(1,1).value = '销售/月份' ;sh.cell(1,2).value = '毛利润合计' ; sh.cell(1,3).value = '计提利润合计'

    border = Border(right=Side(style='thick'),bottom=Side(style='thick'))
    font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
    pattern_fill = PatternFill(fill_type="solid",fgColor="c6e2ff")
    sh.cell(1,1).border = border;sh.cell(1,2).border = border;sh.cell(1,3).border = border
    sh.cell(1,1).font = font;sh.cell(1,2).font = font;sh.cell(1,3).font = font
    sh.cell(1,1).fill = pattern_fill;sh.cell(1,2).fill = pattern_fill;sh.cell(1,3).fill = pattern_fill

    #内容部分
    row = 1
    for name,dct in info.items():
        row += 1
        sh.cell(row,1).value = name; 
        sh.cell(row,2).value = fsum(dct.values())
        sh.cell(row,3).value = ssum(dct.values())
        border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
        font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
        pattern_fill = PatternFill(fill_type="solid",fgColor="ffdead")
        sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
        sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font
        sh.cell(row,1).fill = pattern_fill;sh.cell(row,2).fill = pattern_fill;sh.cell(row,3).fill = pattern_fill

        for month, valpair in info[name].items():
            row+=1
            sh.cell(row,1).value = month; 
            sh.cell(row,2).value = valpair[0] ; sh.cell(row,3).value = valpair[1]
            border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
            font = Font(name="等线",size=8,bold=False,italic=False,color="000000")
            sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
            sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font

    wb.save(dest+'.xlsx')
    print('zhuijia')

def write_table2(dest, sheetname, list_of_business):
    
    wb = openpyxl.load_workbook(dest+'.xlsx')
    wb.create_sheet(sheetname)
    sh = wb[sheetname]
    sh.column_dimensions['A'].width = 12.0
    sh.column_dimensions['B'].width = 12.0
    sh.column_dimensions['C'].width = 12.0
    info = {'01':[0,0],'02':[0,0],'03':[0,0],'04':[0,0],'05':[0,0],'06':[0,0],\
        '07':[0,0],'08':[0,0],'09':[0,0],'10':[0,0],'11':[0,0],'12':[0,0]}
    for business in list_of_business:
        if business[2] not in info:
            info[business[2]] = [0,0]
        info[business[2]][0] += getsums(business)[0]
        info[business[2]][1] += getsums(business)[1]
    
    #表头部分
    sh.cell(1,1).value = '月份' ; sh.cell(1,2).value = '毛利润合计' ; sh.cell(1,3).value = '计提利润合计'

    border = Border(right=Side(style='thick'),bottom=Side(style='thick'))
    font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
    pattern_fill = PatternFill(fill_type="solid",fgColor="c6e2ff")
    sh.cell(1,1).border = border;sh.cell(1,2).border = border;sh.cell(1,3).border = border
    sh.cell(1,1).font = font;sh.cell(1,2).font = font;sh.cell(1,3).font = font
    sh.cell(1,1).fill = pattern_fill;sh.cell(1,2).fill = pattern_fill;sh.cell(1,3).fill = pattern_fill

    #内容部分
    row = 1
    for name,valpair in info.items():
        row += 1
        sh.cell(row,1).value = name;  
        sh.cell(row,2).value = valpair[0] ; sh.cell(row,3).value = valpair[1]
        border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
        font = Font(name="等线",size=8,bold=False,italic=False,color="000000")
        sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
        sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font

    row+=1
    sh.cell(row,1).value = '合计'

    sh.cell(row,2).value = nofr_sum(sh,2,row-1)
    sh.cell(row,3).value = nofr_sum(sh,3,row-1)
    border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
    font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
    pattern_fill = PatternFill(fill_type="solid",fgColor="9f79ee")
    sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
    sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font
    sh.cell(row,1).fill = pattern_fill;sh.cell(row,2).fill = pattern_fill;sh.cell(row,3).fill = pattern_fill
    wb.save(dest+'.xlsx')
    print('zhuijia')

def write_table3(dest, sheetname, list_of_business):
    
    wb = openpyxl.load_workbook(dest+'.xlsx')
    wb.create_sheet(sheetname)
    sh = wb[sheetname]
    sh.column_dimensions['A'].width = 12.0
    sh.column_dimensions['B'].width = 12.0
    sh.column_dimensions['C'].width = 12.0
    info = {}
    for business in list_of_business:
        if business[0] not in info:
            info[business[0]] = {}
        if business[1] not in info[business[0]]:
            info[business[0]][business[1]] = [0,0]
        info[business[0]][business[1]][0] += getsums(business)[0]
        info[business[0]][business[1]][1] += getsums(business)[1]
    
    #表头部分
    sh.cell(1,1).value = '销售/客户' ;sh.cell(1,2).value = '毛利润合计' ; sh.cell(1,3).value = '计提利润合计'

    border = Border(right=Side(style='thick'),bottom=Side(style='thick'))
    font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
    pattern_fill = PatternFill(fill_type="solid",fgColor="c6e2ff")
    sh.cell(1,1).border = border;sh.cell(1,2).border = border;sh.cell(1,3).border = border
    sh.cell(1,1).font = font;sh.cell(1,2).font = font;sh.cell(1,3).font = font
    sh.cell(1,1).fill = pattern_fill;sh.cell(1,2).fill = pattern_fill;sh.cell(1,3).fill = pattern_fill

    #内容部分
    row = 1
    for name,dct in info.items():
        row += 1
        sh.cell(row,1).value = name; 
        sh.cell(row,2).value = fsum(dct.values())
        sh.cell(row,3).value = ssum(dct.values())
        border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
        font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
        pattern_fill = PatternFill(fill_type="solid",fgColor="ffdead")
        sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
        sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font
        sh.cell(row,1).fill = pattern_fill;sh.cell(row,2).fill = pattern_fill;sh.cell(row,3).fill = pattern_fill
        for month, valpair in info[name].items():
            row+=1
            sh.cell(row,1).value = month; 
            sh.cell(row,2).value = valpair[0] ; sh.cell(row,3).value = valpair[1]
            border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
            font = Font(name="等线",size=8,bold=False,italic=False,color="000000")
            sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
            sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font

    wb.save(dest+'.xlsx')
    print('zhuijia')

def write_table4(dest, sheetname, list_of_business):
    
    wb = openpyxl.load_workbook(dest+'.xlsx')
    wb.create_sheet(sheetname)
    sh = wb[sheetname]
    sh.column_dimensions['A'].width = 12.0
    sh.column_dimensions['B'].width = 12.0
    sh.column_dimensions['C'].width = 12.0
    info = {}
    for business in list_of_business:
        if business[1] not in info:
            info[business[1]] = [0,0]
        info[business[1]][0] += getsums(business)[0]
        info[business[1]][1] += getsums(business)[1]
    
    #表头部分
    sh.cell(1,1).value = '客户' ; sh.cell(1,2).value = '毛利润合计' ; sh.cell(1,3).value = '计提利润合计'

    border = Border(right=Side(style='thick'),bottom=Side(style='thick'))
    font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
    pattern_fill = PatternFill(fill_type="solid",fgColor="c6e2ff")
    sh.cell(1,1).border = border;sh.cell(1,2).border = border;sh.cell(1,3).border = border
    sh.cell(1,1).font = font;sh.cell(1,2).font = font;sh.cell(1,3).font = font
    sh.cell(1,1).fill = pattern_fill;sh.cell(1,2).fill = pattern_fill;sh.cell(1,3).fill = pattern_fill

    #内容部分
    row = 1
    for name,valpair in info.items():
        row += 1
        sh.cell(row,1).value = name;  
        sh.cell(row,2).value = valpair[0] ; sh.cell(row,3).value = valpair[1]
        border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
        font = Font(name="等线",size=8,bold=False,italic=False,color="000000")
        sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
        sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font
    row+=1
    sh.cell(row,1).value = '合计'

    sh.cell(row,2).value = nofr_sum(sh,2,row-1)
    sh.cell(row,3).value = nofr_sum(sh,3,row-1)
    border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
    font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
    pattern_fill = PatternFill(fill_type="solid",fgColor="9f79ee")
    sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
    sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font
    sh.cell(row,1).fill = pattern_fill;sh.cell(row,2).fill = pattern_fill;sh.cell(row,3).fill = pattern_fill
    wb.save(dest+'.xlsx')
    print('zhuijia')

def write_table5(dest, sheetname, list_of_business):
    
    wb = openpyxl.load_workbook(dest+'.xlsx')
    wb.create_sheet(sheetname)
    sh = wb[sheetname]
    sh.column_dimensions['A'].width = 12.0
    sh.column_dimensions['B'].width = 12.0
    sh.column_dimensions['C'].width = 12.0
    info = {}
    for business in list_of_business:
        if business[0] not in info:
            info[business[0]] = [[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0]]

        info[business[0]][0][0] += business[3][0];info[business[0]][1][0] += business[3][1]*business[3][0]
        info[business[0]][0][1] += business[4][0];info[business[0]][1][1] += business[4][1]*business[4][0]
        info[business[0]][0][2] += business[5][0];info[business[0]][1][2] += business[5][1]*business[5][0]
        info[business[0]][0][3] += business[6][0];info[business[0]][1][3] += business[6][1]*business[6][0]
        info[business[0]][0][4] += business[7][0];info[business[0]][1][4] += business[7][1]*business[7][0]
        info[business[0]][0][5] += business[8][0];info[business[0]][1][5] += business[8][1]*business[8][0]
        info[business[0]][0][6] += business[9][0];info[business[0]][1][6] += business[9][1]*business[9][0]
        info[business[0]][0][7] += business[10][0];info[business[0]][1][7] += business[10][1]*business[10][0]
        info[business[0]][0][8] += business[11][0];info[business[0]][1][8] += business[11][1]*business[11][0]
        info[business[0]][0][9] += business[12][0];info[business[0]][1][9] += business[12][1]*business[12][0]

    
    #表头部分
    sh.cell(1,1).value = '销售/业务' ;sh.cell(1,2).value = '毛利润合计' ; sh.cell(1,3).value = '计提利润合计'

    border = Border(right=Side(style='thick'),bottom=Side(style='thick'))
    font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
    pattern_fill = PatternFill(fill_type="solid",fgColor="c6e2ff")
    sh.cell(1,1).border = border;sh.cell(1,2).border = border;sh.cell(1,3).border = border
    sh.cell(1,1).font = font;sh.cell(1,2).font = font;sh.cell(1,3).font = font
    sh.cell(1,1).fill = pattern_fill;sh.cell(1,2).fill = pattern_fill;sh.cell(1,3).fill = pattern_fill

    #内容部分
    row = 1
    for name,vals in info.items():
        row += 1
        sh.cell(row,1).value = name; 
        sh.cell(row,2).value = sum(vals[0])
        sh.cell(row,3).value = sum(vals[1])
        border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
        font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
        pattern_fill = PatternFill(fill_type="solid",fgColor="ffdead")
        sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
        sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font
        sh.cell(row,1).fill = pattern_fill;sh.cell(row,2).fill = pattern_fill;sh.cell(row,3).fill = pattern_fill
        
        
        
        yewu = ['空运出口','空运进口','空运关务','海运出口','海运进口','海运关务','快运','铁路出口','铁路进口','国内陆运']

        for i in range(10):
            row+=1
            sh.cell(row,1).value = yewu[i]; 
            sh.cell(row,2).value = vals[0][i] ; sh.cell(row,3).value = vals[1][i]
            border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
            font = Font(name="等线",size=8,bold=False,italic=False,color="000000")
            sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
            sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font
            

    wb.save(dest+'.xlsx')
    print('zhuijia')

def write_table6(dest, sheetname, list_of_business):
    
    wb = openpyxl.load_workbook(dest+'.xlsx')
    wb.create_sheet(sheetname)
    sh = wb[sheetname]
    sh.column_dimensions['A'].width = 12.0
    sh.column_dimensions['B'].width = 12.0
    sh.column_dimensions['C'].width = 12.0
    info = {}
    for business in list_of_business:
        if business[0] not in info:
            info[business[0]] = [[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0]]

        info[business[0]][0][0] += business[3][0];info[business[0]][1][0] += business[3][1]*business[3][0]
        info[business[0]][0][1] += business[4][0];info[business[0]][1][1] += business[4][1]*business[4][0]
        info[business[0]][0][2] += business[5][0];info[business[0]][1][2] += business[5][1]*business[5][0]
        info[business[0]][0][3] += business[6][0];info[business[0]][1][3] += business[6][1]*business[6][0]
        info[business[0]][0][4] += business[7][0];info[business[0]][1][4] += business[7][1]*business[7][0]
        info[business[0]][0][5] += business[8][0];info[business[0]][1][5] += business[8][1]*business[8][0]
        info[business[0]][0][6] += business[9][0];info[business[0]][1][6] += business[9][1]*business[9][0]
        info[business[0]][0][7] += business[10][0];info[business[0]][1][7] += business[10][1]*business[10][0]
        info[business[0]][0][8] += business[11][0];info[business[0]][1][8] += business[11][1]*business[11][0]
        info[business[0]][0][9] += business[12][0];info[business[0]][1][9] += business[12][1]*business[12][0]

    
    #表头部分
    sh.cell(1,1).value = '业务' ;sh.cell(1,2).value = '毛利润合计' ; sh.cell(1,3).value = '计提利润合计'

    border = Border(right=Side(style='thick'),bottom=Side(style='thick'))
    font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
    pattern_fill = PatternFill(fill_type="solid",fgColor="c6e2ff")
    sh.cell(1,1).border = border;sh.cell(1,2).border = border;sh.cell(1,3).border = border
    sh.cell(1,1).font = font;sh.cell(1,2).font = font;sh.cell(1,3).font = font
    sh.cell(1,1).fill = pattern_fill;sh.cell(1,2).fill = pattern_fill;sh.cell(1,3).fill = pattern_fill

    fsum = [0,0,0,0,0,0,0,0,0,0];ssum = [0,0,0,0,0,0,0,0,0,0]
    for name, vals in info.items():
        for i in range(10):
            fsum[i] += vals[0][i]
            ssum[i] += vals[1][i]
    
    #内容部分
    row = 1
    
    yewu = ['空运出口','空运进口','空运关务','海运出口','海运进口','海运关务','快运','铁路出口','铁路进口','国内陆运',]

    for i in range(10):
        row+=1
        sh.cell(row,1).value = yewu[i]; 
        sh.cell(row,2).value = fsum[i] ; sh.cell(row,3).value = ssum[i]
        border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
        font = Font(name="等线",size=8,bold=False,italic=False,color="000000")
        sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
        sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font
            
    row+=1
    sh.cell(row,1).value = '合计'

    sh.cell(row,2).value = nofr_sum(sh,2,row-1)
    sh.cell(row,3).value = nofr_sum(sh,3,row-1)
    border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
    font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
    pattern_fill = PatternFill(fill_type="solid",fgColor="9f79ee")
    sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
    sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font
    sh.cell(row,1).fill = pattern_fill;sh.cell(row,2).fill = pattern_fill;sh.cell(row,3).fill = pattern_fill

    wb.save(dest+'.xlsx')
    print('zhuijia')

def write_table7(dest, sheetname, list_of_business):
    
    wb = openpyxl.load_workbook(dest+'.xlsx')
    wb.create_sheet(sheetname)
    sh = wb[sheetname]
    sh.column_dimensions['A'].width = 12.0
    sh.column_dimensions['B'].width = 12.0
    sh.column_dimensions['C'].width = 12.0
    info = {}
    for business in list_of_business:
        if business[2] not in info:
            info[business[2]] = [[0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0]]

        info[business[2]][0][0] += business[3][0];info[business[2]][1][0] += business[3][1]*business[3][0]
        info[business[2]][0][1] += business[4][0];info[business[2]][1][1] += business[4][1]*business[4][0]
        info[business[2]][0][2] += business[5][0];info[business[2]][1][2] += business[5][1]*business[5][0]
        info[business[2]][0][3] += business[6][0];info[business[2]][1][3] += business[6][1]*business[6][0]
        info[business[2]][0][4] += business[7][0];info[business[2]][1][4] += business[7][1]*business[7][0]
        info[business[2]][0][5] += business[8][0];info[business[2]][1][5] += business[8][1]*business[8][0]
        info[business[2]][0][6] += business[9][0];info[business[2]][1][6] += business[9][1]*business[9][0]
        info[business[2]][0][7] += business[10][0];info[business[2]][1][7] += business[10][1]*business[10][0]
        info[business[2]][0][8] += business[11][0];info[business[2]][1][8] += business[11][1]*business[11][0]
        info[business[2]][0][9] += business[12][0];info[business[2]][1][9] += business[12][1]*business[12][0]

    
    #表头部分
    sh.cell(1,1).value = '月份' ;sh.cell(1,2).value = '毛利润合计' ; sh.cell(1,3).value = '计提利润合计'

    border = Border(right=Side(style='thick'),bottom=Side(style='thick'))
    font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
    pattern_fill = PatternFill(fill_type="solid",fgColor="c6e2ff")
    sh.cell(1,1).border = border;sh.cell(1,2).border = border;sh.cell(1,3).border = border
    sh.cell(1,1).font = font;sh.cell(1,2).font = font;sh.cell(1,3).font = font
    sh.cell(1,1).fill = pattern_fill;sh.cell(1,2).fill = pattern_fill;sh.cell(1,3).fill = pattern_fill

    #内容部分
    row = 1
    for name,vals in info.items():
        row += 1
        sh.cell(row,1).value = name; 
        sh.cell(row,2).value = sum(vals[0])
        sh.cell(row,3).value = sum(vals[1])
        border = Border(right=Side(style='medium'),bottom=Side(style='medium'))
        font = Font(name="等线",size=8,bold=True,italic=False,color="000000")
        pattern_fill = PatternFill(fill_type="solid",fgColor="ffdead")
        sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
        sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font
        sh.cell(row,1).fill = pattern_fill;sh.cell(row,2).fill = pattern_fill;sh.cell(row,3).fill = pattern_fill
        
        
        
        yewu = ['空运出口','空运进口','空运关务','海运出口','海运进口','海运关务','快运','铁路出口','铁路进口','国内陆运']

        for i in range(10):
            row+=1
            sh.cell(row,1).value = yewu[i]; 
            sh.cell(row,2).value = vals[0][i] ; sh.cell(row,3).value = vals[1][i]
            border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
            font = Font(name="等线",size=8,bold=False,italic=False,color="000000")
            sh.cell(row,1).border = border;sh.cell(row,2).border = border;sh.cell(row,3).border = border
            sh.cell(row,1).font = font;sh.cell(row,2).font = font;sh.cell(row,3).font = font
            

    wb.save(dest+'.xlsx')
    print('zhuijia')

'''
 # 导出类型: 仅人(0), 人--月份(1), 仅月份(2), 人--客户(3), 仅客户(4), 人--业务(5), 仅业务(6)
        afunc.write_table0(dest, sheetname == '仅人', list_business)
        afunc.write_table1(dest, sheetname == '人--月份', list_business)
        afunc.write_table2(dest, sheetname == '仅月份', list_business)
        afunc.write_table3(dest, sheetname == '人--客户', list_business)
        afunc.write_table4(dest, sheetname == '仅客户', list_business)
        afunc.write_table5(dest, sheetname == '人--业务', list_business)
        afunc.write_table6(dest, sheetname == '仅业务', list_business)

C:/Users/Administrator/Desktop/副本狄斯洋-2310-销售利润统计表.xlsx ; 狄斯洋
C:/Users/Administrator/Desktop/副本宋欣欣-2311-销售利润统计表.xlsx ; 宋欣欣
C:/Users/Administrator/Desktop/副本张松-2311-销售利润统计表.xlsx ; 张松
C:/Users/Administrator/Desktop/副本朱琳-2311-销售利润统计表.xlsx ; 朱琳
'''

if __name__ == '__main__':
    fn = 'C:/Users/Administrator/Desktop/yxm_tablecombine/合并表格.xlsx'
    list_business = write_table_from_filenames('12345',[fn],['汇总表格'])

    write_table0('12345','仅人',list_business)
    write_table1('12345', '人--月份', list_business)
    write_table2('12345', '仅月份', list_business)
    write_table3('12345', '人--客户', list_business)
    write_table4('12345', '仅客户', list_business)
    write_table5('12345', '人--业务', list_business)
    write_table6('12345', '仅业务', list_business)
